import argparse
import json
import re
import zipfile
from datetime import date, datetime
from pathlib import Path
from typing import Any, Dict, List, Optional, Tuple
import xml.etree.ElementTree as ET


WB_NS = {"m": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
REL_NS = {"r": "http://schemas.openxmlformats.org/package/2006/relationships"}
RID_ATTR = "{http://schemas.openxmlformats.org/officeDocument/2006/relationships}id"
CELL_REF_RE = re.compile(r"^([A-Z]+)([0-9]+)$")


def _col_to_idx(col: str) -> int:
    n = 0
    for ch in col:
        n = n * 26 + (ord(ch) - ord("A") + 1)
    return n - 1


def _idx_to_col(idx: int) -> str:
    n = idx + 1
    out = []
    while n:
        n, r = divmod(n - 1, 26)
        out.append(chr(ord("A") + r))
    return "".join(reversed(out))


def _to_jsonable(v: Any) -> Any:
    if v is None:
        return None
    if isinstance(v, (str, int, float, bool)):
        return v
    if isinstance(v, (datetime, date)):
        return v.isoformat()
    return str(v)


def _trim_trailing_empty(row: List[Any]) -> List[Any]:
    i = len(row)
    while i > 0 and (row[i - 1] is None or row[i - 1] == ""):
        i -= 1
    return row[:i]


def _load_shared_strings(z: zipfile.ZipFile) -> List[str]:
    if "xl/sharedStrings.xml" not in z.namelist():
        return []
    root = ET.fromstring(z.read("xl/sharedStrings.xml"))
    out: List[str] = []
    for si in root.findall("m:si", WB_NS):
        t = si.find("m:t", WB_NS)
        if t is not None and t.text is not None:
            out.append(t.text)
            continue
        parts: List[str] = []
        for r in si.findall("m:r", WB_NS):
            tt = r.find("m:t", WB_NS)
            if tt is not None and tt.text is not None:
                parts.append(tt.text)
        out.append("".join(parts))
    return out


def _read_sheet_grid(
    z: zipfile.ZipFile,
    sheet_path: str,
    shared_strings: List[str],
    max_rows: Optional[int] = None,
    max_cols: Optional[int] = None,
    drop_empty_rows: bool = True,
) -> Tuple[List[List[Any]], int]:
    root = ET.fromstring(z.read(sheet_path))
    sheet_data = root.find("m:sheetData", WB_NS)
    if sheet_data is None:
        return [], 0

    rows_out: List[List[Any]] = []
    max_col_seen = 0

    for row in sheet_data.findall("m:row", WB_NS):
        r_attr = row.attrib.get("r")
        rnum = int(r_attr) if r_attr and r_attr.isdigit() else None
        if max_rows is not None and rnum is not None and rnum > max_rows:
            break

        cells: Dict[int, Any] = {}

        for c in row.findall("m:c", WB_NS):
            ref = c.attrib.get("r")
            if not ref:
                continue
            m = CELL_REF_RE.match(ref)
            if not m:
                continue
            col_letters = m.group(1)
            col_idx = _col_to_idx(col_letters)
            if max_cols is not None and col_idx >= max_cols:
                continue

            t = c.attrib.get("t")
            v = c.find("m:v", WB_NS)
            is_inline_str = t == "inlineStr"

            value: Any
            if is_inline_str:
                is_el = c.find("m:is", WB_NS)
                tt = is_el.find("m:t", WB_NS) if is_el is not None else None
                value = tt.text if (tt is not None and tt.text is not None) else ""
            elif v is None or v.text is None:
                f = c.find("m:f", WB_NS)
                value = f.text if (f is not None and f.text is not None) else ""
            else:
                raw = v.text
                if t == "s":
                    try:
                        value = shared_strings[int(raw)]
                    except Exception:
                        value = raw
                elif t == "b":
                    value = raw == "1"
                else:
                    value = raw

            cells[col_idx] = value
            if col_idx + 1 > max_col_seen:
                max_col_seen = col_idx + 1

        if not cells:
            if not drop_empty_rows:
                rows_out.append([])
            continue

        row_len = (max(cells.keys()) + 1) if cells else 0
        row_vals: List[Any] = [""] * row_len
        for idx, val in cells.items():
            if idx < len(row_vals):
                row_vals[idx] = val

        rows_out.append(row_vals)

    return rows_out, max_col_seen


def _zip_load_workbook(
    path: Path,
    max_rows: Optional[int],
    max_cols: Optional[int],
    drop_empty_rows: bool,
) -> Dict[str, Any]:
    with zipfile.ZipFile(path) as z:
        names = set(z.namelist())
        if "xl/workbook.xml" not in names:
            raise ValueError("Missing xl/workbook.xml")

        shared_strings = _load_shared_strings(z)

        rid_to_target: Dict[str, str] = {}
        if "xl/_rels/workbook.xml.rels" in names:
            rels = ET.fromstring(z.read("xl/_rels/workbook.xml.rels"))
            for rel in rels.findall("r:Relationship", REL_NS):
                rid = rel.attrib.get("Id")
                target = rel.attrib.get("Target")
                if rid and target:
                    rid_to_target[rid] = target

        wb = ET.fromstring(z.read("xl/workbook.xml"))
        sheets: List[Dict[str, Any]] = []
        for sh in wb.findall("m:sheets/m:sheet", WB_NS):
            name = sh.attrib.get("name")
            rid = sh.attrib.get(RID_ATTR)
            target = rid_to_target.get(rid) if rid else None
            sheet_path = ("xl/" + target.lstrip("/")) if target else None

            grid: List[List[Any]] = []
            max_col_seen = 0
            if sheet_path and sheet_path in names:
                grid, max_col_seen = _read_sheet_grid(
                    z,
                    sheet_path,
                    shared_strings,
                    max_rows=max_rows,
                    max_cols=max_cols,
                    drop_empty_rows=drop_empty_rows,
                )

            sheets.append(
                {
                    "name": name,
                    "rid": rid,
                    "target": target,
                    "grid": grid,
                    "max_col_seen": max_col_seen,
                }
            )

    return {
        "engine": "xlsx-xml",
        "source": str(path),
        "extracted_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "sheets": sheets,
    }


def _openpyxl_load_workbook(
    path: Path,
    max_rows: Optional[int],
    max_cols: Optional[int],
    drop_empty_rows: bool,
) -> Dict[str, Any]:
    import openpyxl  # type: ignore

    wb = openpyxl.load_workbook(path, data_only=False, read_only=True)

    sheets: List[Dict[str, Any]] = []
    for ws in wb.worksheets:
        grid: List[List[Any]] = []
        max_col_seen = 0

        row_limit = max_rows if max_rows is not None else ws.max_row
        col_limit = max_cols if max_cols is not None else ws.max_column

        for r in range(1, row_limit + 1):
            row_vals: List[Any] = []
            any_value = False
            for c in range(1, col_limit + 1):
                v = ws.cell(row=r, column=c).value
                vj = _to_jsonable(v)
                if vj not in (None, ""):
                    any_value = True
                row_vals.append(vj if vj is not None else "")

            if drop_empty_rows and not any_value:
                continue

            trimmed = _trim_trailing_empty(row_vals)
            if len(trimmed) > max_col_seen:
                max_col_seen = len(trimmed)
            grid.append(trimmed)

        sheets.append(
            {
                "name": ws.title,
                "grid": grid,
                "max_col_seen": max_col_seen,
            }
        )

    return {
        "engine": "openpyxl",
        "source": str(path),
        "extracted_at": datetime.utcnow().isoformat(timespec="seconds") + "Z",
        "sheets": sheets,
    }


def load_workbook_to_dict(
    path: Path,
    max_rows: Optional[int],
    max_cols: Optional[int],
    drop_empty_rows: bool,
) -> Dict[str, Any]:
    try:
        return _openpyxl_load_workbook(path, max_rows=max_rows, max_cols=max_cols, drop_empty_rows=drop_empty_rows)
    except Exception:
        return _zip_load_workbook(path, max_rows=max_rows, max_cols=max_cols, drop_empty_rows=drop_empty_rows)


def _preview_payload(payload: Dict[str, Any], rows: int, cols: int) -> None:
    sheets = payload.get("sheets") or []
    print(f"engine={payload.get('engine')}")
    print(f"sheets={len(sheets)}")

    for sh in sheets:
        name = sh.get("name")
        grid = sh.get("grid") or []
        print(f"\n[{name}]")
        for r in grid[:rows]:
            r2 = (r + [""] * cols)[:cols]
            print(r2)


def main() -> int:
    parser = argparse.ArgumentParser(prog="excel_to_json")
    sub = parser.add_subparsers(dest="cmd", required=True)

    p_preview = sub.add_parser("preview")
    p_preview.add_argument("input", type=str)
    p_preview.add_argument("--rows", type=int, default=10)
    p_preview.add_argument("--cols", type=int, default=12)
    p_preview.add_argument("--max-rows", type=int, default=None)
    p_preview.add_argument("--max-cols", type=int, default=None)
    p_preview.add_argument("--keep-empty-rows", action="store_true")

    p_export = sub.add_parser("export")
    p_export.add_argument("input", type=str)
    p_export.add_argument("--out", type=str, default=None)
    p_export.add_argument("--max-rows", type=int, default=None)
    p_export.add_argument("--max-cols", type=int, default=None)
    p_export.add_argument("--keep-empty-rows", action="store_true")

    args = parser.parse_args()

    input_path = Path(args.input)
    if not input_path.exists():
        raise SystemExit(f"Input not found: {input_path}")

    drop_empty_rows = not getattr(args, "keep_empty_rows", False)

    payload = load_workbook_to_dict(
        input_path,
        max_rows=getattr(args, "max_rows", None),
        max_cols=getattr(args, "max_cols", None),
        drop_empty_rows=drop_empty_rows,
    )

    if args.cmd == "preview":
        _preview_payload(payload, rows=args.rows, cols=args.cols)
        return 0

    out_path: Path
    if args.out:
        out_path = Path(args.out)
    else:
        out_dir = Path("data") / "excel_exports"
        out_dir.mkdir(parents=True, exist_ok=True)
        out_path = out_dir / (input_path.stem + ".json")

    out_path.parent.mkdir(parents=True, exist_ok=True)
    with out_path.open("w", encoding="utf-8") as f:
        json.dump(payload, f, ensure_ascii=False, indent=2)

    print(str(out_path))
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
